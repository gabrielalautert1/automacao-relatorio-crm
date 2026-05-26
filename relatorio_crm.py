import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

arquivo_entrada = input("Nome do arquivo Excel (ex: organico13.03.xlsx): ")

try:
    df_ads = pd.read_excel(arquivo_entrada)
except FileNotFoundError:
    print(f"Arquivo '{arquivo_entrada}' não encontrado.")
    exit()

df_ads.columns = [c.strip() for c in df_ads.columns]

wb = Workbook()
ws = wb.active
ws.title = "Resumo Detalhado"

bold_font = Font(bold=True)
fill_perdido = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
fill_ativo = PatternFill(start_color="F0E5FF", end_color="F0E5FF", fill_type="solid")
fill_ganho = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
fill_resumo = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
centralizado = Alignment(vertical="center")

cadencias = df_ads["Cadência"].dropna().unique()
linha = 1

for cadencia in cadencias:
    df_cad = df_ads[df_ads["Cadência"] == cadencia]

    total = len(df_cad)
    ativos = len(df_cad[df_cad["Status"] == "Ativo"])
    ganhos = len(df_cad[df_cad["Status"] == "Ganho"])
    perdidos = len(df_cad[df_cad["Status"] == "Perdido"])

    ws[f"A{linha}"] = f"[{cadencia}] — {total} Leads"
    ws[f"A{linha}"].font = bold_font
    linha += 1

    ws[f"A{linha}"] = f"{perdidos} Perdidos"
    ws[f"A{linha}"].fill = fill_perdido
    linha += 1

    df_perdidos = df_cad[df_cad["Status"] == "Perdido"]
    motivos = df_perdidos["Motivo de perda"].dropna().unique()

    if len(motivos) > 0:
        ws[f"A{linha}"] = "Motivos:"
        linha += 1
